import os
from pathlib import Path
from collections import defaultdict
import hashlib
import pandas as pd
from datetime import datetime


class VideoDatasetAnalyzer:
    """
    Comprehensive tool for video dataset analysis and validation.
    Combines dataset structure analysis with data leakage detection.
    """
    
    def __init__(self, split_dataset_path, original_dataset_path=None):
        self.split_path = Path(split_dataset_path)
        self.original_path = Path(original_dataset_path) if original_dataset_path else None
        self.video_extensions = {'.mp4', '.avi', '.mov', '.mkv', '.MP4', '.AVI', '.MOV', '.MKV'}
        
        # Category mapping: split names (with underscores) to original names (with spaces)
        self.category_mapping = {
            'Natural_Content': 'Natural Content',
            'Flat_Content': 'Flat Content',
            'Animation': 'Animation',
            'Gaming': 'Gaming'
        }
        
        # Complete category structure
        self.category_structure = {
            'Animation': [
                'Animation', 'Bleach', 'Cartoon', 'Dragon Ball',
                'Lego minifigure', 'Naruto', 'One Piece', 
                'Sonic the Hedgehog', 'The Walt Disney Company'
            ],
            'Gaming': [
                'Action-adventure game', 'Battlefield', 'Call of Duty',
                'FIFA 15', 'Games', 'Grand Theft Auto', 'Grand Theft Auto V',
                'League of Legends', 'Minecraft', 'RuneScape', 
                'Video game', 'World of Warcraft'
            ],
            'Natural_Content': [
                'Animal', 'Bird', 'Cat', 'Chicken', 'Dog', 'Farm', 'Fish',
                'Fishing', 'Garden', 'Horse', 'Nature', 'Outdoor recreation',
                'Pet', 'Plant', 'Tree', 'Wildlife'
            ],
            'Flat_Content': [
                'Chart', 'Illustration', 'Logo', 'Map', 'Poster',
                'Screencast', 'Text', 'Typography', 'Website'
            ]
        }
    
    # ==================== DATASET STRUCTURE ANALYSIS ====================
    
    def analyze_original_dataset(self):
        """Analyze original dataset structure: original_path/category/videos/subcategory"""
        original_counts = {}
        
        if not self.original_path or not self.original_path.exists():
            if self.original_path:
                print(f"Warning: Original dataset path {self.original_path} does not exist!")
            return original_counts
        
        print(f"\nAnalyzing original dataset at: {self.original_path}")
        
        for main_category, subcategories in self.category_structure.items():
            original_category_name = self.category_mapping.get(main_category, main_category)
            
            for subcategory in subcategories:
                subcategory_path = self.original_path / original_category_name / 'videos' / subcategory
                
                video_count = 0
                if subcategory_path.exists():
                    video_files = [
                        f for f in subcategory_path.iterdir()
                        if f.is_file() and f.suffix.lower() in self.video_extensions
                    ]
                    video_count = len(video_files)
                
                original_counts[(main_category, subcategory)] = video_count
        
        return original_counts
    
    def analyze_split_dataset(self):
        """Analyze train/val/test split structure and counts"""
        splits = ['train', 'val', 'test']
        data_rows = []
        
        original_counts = self.analyze_original_dataset() if self.original_path else {}
        
        print(f"\nAnalyzing split dataset at: {self.split_path}")
        print("=" * 80)
        
        for main_category, subcategories in self.category_structure.items():
            print(f"\nProcessing {main_category}...")
            
            for subcategory in subcategories:
                row_data = {
                    'Main Category': main_category,
                    'Subcategory': subcategory,
                    'Original': original_counts.get((main_category, subcategory), 0),
                    'Train': 0,
                    'Validation': 0,
                    'Test': 0,
                    'Total Split': 0
                }
                
                for split in splits:
                    split_dir = self.split_path / split
                    if not split_dir.exists():
                        continue
                    
                    subcategory_path = split_dir / main_category / subcategory
                    
                    video_count = 0
                    if subcategory_path.exists():
                        video_files = [
                            f for f in subcategory_path.iterdir()
                            if f.is_file() and f.suffix.lower() in self.video_extensions
                        ]
                        video_count = len(video_files)
                    
                    split_column_map = {
                        'train': 'Train',
                        'val': 'Validation',
                        'test': 'Test'
                    }
                    
                    row_data[split_column_map[split]] = video_count
                    row_data['Total Split'] += video_count
                
                data_rows.append(row_data)
                
                if self.original_path:
                    print(f"  {subcategory}: Original={row_data['Original']}, "
                          f"Train={row_data['Train']}, Val={row_data['Validation']}, "
                          f"Test={row_data['Test']}, Total={row_data['Total Split']}")
                else:
                    print(f"  {subcategory}: Train={row_data['Train']}, "
                          f"Val={row_data['Validation']}, Test={row_data['Test']}, "
                          f"Total={row_data['Total Split']}")
        
        return pd.DataFrame(data_rows)
    
    def export_structure_analysis(self, output_dir='results', output_file='dataset_complete_analysis.xlsx'):
        """Export dataset structure analysis to Excel"""
        df = self.analyze_split_dataset()
        
        if df.empty:
            print("No data to export!")
            return None
        
        # Create output directory if it doesn't exist
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        output_filepath = output_path / output_file
        
        # Sort by category and subcategory
        df = df.sort_values(['Main Category', 'Subcategory'], ascending=[True, True])
        
        # Create category totals
        final_rows = []
        for category in self.category_structure.keys():
            cat_rows = df[df['Main Category'] == category].to_dict('records')
            final_rows.extend(cat_rows)
            
            # Add total row
            cat_df = df[df['Main Category'] == category]
            total_row = {
                'Main Category': '',
                'Subcategory': f'Total: {len(cat_df)}',
                'Original': cat_df['Original'].sum(),
                'Train': cat_df['Train'].sum(),
                'Validation': cat_df['Validation'].sum(),
                'Test': cat_df['Test'].sum(),
                'Total Split': cat_df['Total Split'].sum()
            }
            final_rows.append(total_row)
            
            # Add blank separator
            if category != list(self.category_structure.keys())[-1]:
                final_rows.append({k: '' for k in total_row.keys()})
        
        final_df = pd.DataFrame(final_rows)
        
        # Export to Excel
        try:
            with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
                final_df.to_excel(writer, sheet_name='Dataset Analysis', index=False)
                
                from openpyxl.styles import Font, PatternFill, Alignment
                
                worksheet = writer.sheets['Dataset Analysis']
                
                # Format header
                header_fill = PatternFill(start_color='B4C7DC', end_color='B4C7DC', fill_type='solid')
                header_font = Font(bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 20
                worksheet.column_dimensions['B'].width = 30
                for col in ['C', 'D', 'E', 'F', 'G']:
                    worksheet.column_dimensions[col].width = 12
                
                # Format total rows
                total_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                total_font = Font(bold=True)
                
                for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    cell_value = worksheet[f'B{idx}'].value
                    if cell_value and str(cell_value).startswith('Total:'):
                        for cell in row:
                            cell.fill = total_fill
                            cell.font = total_font
                
                # Center align numeric columns
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=7):
                    for cell in row:
                        if cell.value != '':
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"\n✓ Structure analysis exported to: {output_filepath}")
            print(f"  Location: {output_filepath.absolute()}")
            self._print_summary_statistics(df)
            return final_df
            
        except Exception as e:
            print(f"\n✗ Error exporting to Excel: {e}")
            csv_file = output_filepath.with_suffix('.csv')
            final_df.to_csv(csv_file, index=False)
            print(f"✓ Exported to CSV instead: {csv_file}")
            return final_df
    
    def _print_summary_statistics(self, df):
        """Print summary statistics"""
        print("\n" + "=" * 80)
        print("SUMMARY STATISTICS")
        print("=" * 80)
        
        if self.original_path:
            total_original = df['Original'].sum()
            print(f"\nOriginal Dataset: {total_original} videos")
        
        total_videos = df['Total Split'].sum()
        total_train = df['Train'].sum()
        total_val = df['Validation'].sum()
        total_test = df['Test'].sum()
        
        print(f"\nSplit Dataset Total: {total_videos} videos")
        print(f"  Train:      {total_train:5} videos ({total_train/total_videos*100:5.1f}%)")
        print(f"  Validation: {total_val:5} videos ({total_val/total_videos*100:5.1f}%)")
        print(f"  Test:       {total_test:5} videos ({total_test/total_videos*100:5.1f}%)")
        
        print("\nCategory Breakdown:")
        for category in self.category_structure.keys():
            cat_df = df[df['Main Category'] == category]
            cat_total = cat_df['Total Split'].sum()
            if self.original_path:
                cat_original = cat_df['Original'].sum()
                print(f"  {category:20}: Original={cat_original:5} | Split={cat_total:5} ({cat_total/total_videos*100:5.1f}%)")
            else:
                print(f"  {category:20}: {cat_total:5} videos ({cat_total/total_videos*100:5.1f}%)")
    
    # ==================== DATA LEAKAGE CHECKING ====================
    
    def get_video_files(self, split):
        """Get all video files from a split"""
        split_path = self.split_path / split
        videos = []
        
        if not split_path.exists():
            print(f"WARNING: {split_path} does not exist!")
            return videos
        
        for ext in self.video_extensions:
            videos.extend(list(split_path.rglob(f"*{ext}")))
        
        return videos
    
    def get_file_hash(self, file_path, chunk_size=8192):
        """Calculate MD5 hash of a file"""
        md5 = hashlib.md5()
        try:
            with open(file_path, 'rb') as f:
                while chunk := f.read(chunk_size):
                    md5.update(chunk)
            return md5.hexdigest()
        except Exception as e:
            print(f"Error hashing {file_path}: {e}")
            return None
    
    def check_filename_duplicates(self, export_excel=True, output_dir='results'):
        """Check for duplicate filenames across different splits"""
        print("\n" + "=" * 70)
        print("CHECKING FOR FILENAME DUPLICATES ACROSS SPLITS")
        print("=" * 70)
        
        train_files = self.get_video_files('train')
        val_files = self.get_video_files('val')
        test_files = self.get_video_files('test')
        
        print(f"\nTotal files found:")
        print(f"  Train: {len(train_files)} videos")
        print(f"  Val:   {len(val_files)} videos")
        print(f"  Test:  {len(test_files)} videos")
        print(f"  TOTAL: {len(train_files) + len(val_files) + len(test_files)} videos")
        
        # Create filename mapping
        filename_map = defaultdict(list)
        
        for file_path in train_files:
            filename_map[file_path.name].append(('train', file_path))
        for file_path in val_files:
            filename_map[file_path.name].append(('val', file_path))
        for file_path in test_files:
            filename_map[file_path.name].append(('test', file_path))
        
        # Find cross-split duplicates
        cross_split_duplicates = {}
        for name, locations in filename_map.items():
            splits_found = set(split for split, _ in locations)
            if len(splits_found) > 1:
                cross_split_duplicates[name] = locations
        
        if not cross_split_duplicates:
            print("\n✓ NO CROSS-SPLIT FILENAME DUPLICATES FOUND!")
            return True
        else:
            print(f"\n✗ FOUND {len(cross_split_duplicates)} FILENAMES IN MULTIPLE SPLITS!")
            
            excel_data = []
            for filename, locations in sorted(cross_split_duplicates.items()):
                splits = sorted(set(split for split, _ in locations))
                print(f"\n  File: {filename}")
                print(f"  Appears in splits: {', '.join(splits)} ⚠")
                
                for split, path in locations:
                    print(f"    [{split}] {path.relative_to(self.split_path)}")
                    
                    file_size = path.stat().st_size if path.exists() else 0
                    file_size_mb = file_size / (1024 * 1024)
                    
                    excel_data.append({
                        'Filename': filename,
                        'Split': split,
                        'Full_Path': str(path),
                        'Relative_Path': str(path.relative_to(self.split_path)),
                        'File_Size_MB': round(file_size_mb, 2),
                        'Splits_Found_In': ', '.join(splits)
                    })
            
            if export_excel and excel_data:
                self._export_duplicates_to_excel(excel_data, 'filename', output_dir)
            
            return False
    
    def check_content_duplicates(self, sample_size=None, export_excel=True, output_dir='results'):
        """Check for duplicate file content across different splits"""
        print("\n" + "=" * 70)
        print("CHECKING FOR CONTENT DUPLICATES ACROSS SPLITS")
        print("=" * 70)
        
        train_files = self.get_video_files('train')
        val_files = self.get_video_files('val')
        test_files = self.get_video_files('test')
        
        if sample_size:
            import random
            random.seed(42)
            train_files = random.sample(train_files, min(sample_size, len(train_files)))
            val_files = random.sample(val_files, min(sample_size, len(val_files)))
            test_files = random.sample(test_files, min(sample_size, len(test_files)))
            print(f"Using sample of {sample_size} files per split")
        
        print(f"\nHashing files...")
        print(f"  Train: {len(train_files)} files")
        print(f"  Val:   {len(val_files)} files")
        print(f"  Test:  {len(test_files)} files")
        
        hash_map = defaultdict(list)
        
        for file_list, split in [(train_files, 'train'), (val_files, 'val'), (test_files, 'test')]:
            print(f"Processing {split} files...")
            for i, file_path in enumerate(file_list, 1):
                file_hash = self.get_file_hash(file_path)
                if file_hash:
                    hash_map[file_hash].append((split, file_path))
                if i % 100 == 0:
                    print(f"  Processed {i}/{len(file_list)} files")
        
        # Find cross-split content duplicates
        cross_split_content_duplicates = {}
        for file_hash, locations in hash_map.items():
            splits_found = set(split for split, _ in locations)
            if len(splits_found) > 1:
                cross_split_content_duplicates[file_hash] = locations
        
        if not cross_split_content_duplicates:
            print("\n✓ NO CROSS-SPLIT CONTENT DUPLICATES FOUND!")
            return True
        else:
            print(f"\n✗ FOUND {len(cross_split_content_duplicates)} IDENTICAL VIDEOS IN MULTIPLE SPLITS!")
            
            excel_data = []
            display_count = min(10, len(cross_split_content_duplicates))
            
            for idx, (file_hash, locations) in enumerate(sorted(cross_split_content_duplicates.items())[:display_count], 1):
                splits = sorted(set(split for split, _ in locations))
                print(f"\n  Group {idx} - Hash: {file_hash[:16]}...")
                print(f"  Appears in splits: {', '.join(splits)} ⚠")
                
                for split, path in locations:
                    print(f"    [{split}] {path.name}")
                    
                    file_size = path.stat().st_size if path.exists() else 0
                    file_size_mb = file_size / (1024 * 1024)
                    
                    excel_data.append({
                        'Filename': path.name,
                        'Split': split,
                        'Full_Path': str(path),
                        'Relative_Path': str(path.relative_to(self.split_path)),
                        'File_Size_MB': round(file_size_mb, 2),
                        'MD5_Hash': file_hash,
                        'Duplicate_Group': f"Group_{idx}",
                        'Splits_Found_In': ', '.join(splits)
                    })
            
            if len(cross_split_content_duplicates) > display_count:
                print(f"\n  ... and {len(cross_split_content_duplicates) - display_count} more")
            
            if export_excel and excel_data:
                self._export_duplicates_to_excel(excel_data, 'content', output_dir)
            
            return False
    
    def check_split_integrity(self):
        """Check if files are properly distributed across splits"""
        print("\n" + "=" * 70)
        print("CHECKING SPLIT INTEGRITY")
        print("=" * 70)
        
        train_files = self.get_video_files('train')
        val_files = self.get_video_files('val')
        test_files = self.get_video_files('test')
        
        total = len(train_files) + len(val_files) + len(test_files)
        
        if total == 0:
            print("\n✗ ERROR: No video files found in any split!")
            return False
        
        train_ratio = len(train_files) / total
        val_ratio = len(val_files) / total
        test_ratio = len(test_files) / total
        
        print(f"\nSplit distribution:")
        print(f"  Train: {len(train_files):5d} files ({train_ratio*100:5.2f}%)")
        print(f"  Val:   {len(val_files):5d} files ({val_ratio*100:5.2f}%)")
        print(f"  Test:  {len(test_files):5d} files ({test_ratio*100:5.2f}%)")
        print(f"  Total: {total:5d} files")
        
        expected_train, expected_val, expected_test = 0.7, 0.2, 0.1
        tolerance = 0.05
        
        print(f"\nExpected ratios (±{tolerance*100}% tolerance):")
        print(f"  Train: {expected_train*100:.1f}% (found: {train_ratio*100:.2f}%)")
        print(f"  Val:   {expected_val*100:.1f}% (found: {val_ratio*100:.2f}%)")
        print(f"  Test:  {expected_test*100:.1f}% (found: {test_ratio*100:.2f}%)")
        
        issues = []
        if abs(train_ratio - expected_train) > tolerance:
            issues.append(f"Train ratio {train_ratio*100:.2f}% deviates from expected {expected_train*100}%")
        if abs(val_ratio - expected_val) > tolerance:
            issues.append(f"Val ratio {val_ratio*100:.2f}% deviates from expected {expected_val*100}%")
        if abs(test_ratio - expected_test) > tolerance:
            issues.append(f"Test ratio {test_ratio*100:.2f}% deviates from expected {expected_test*100}%")
        
        if issues:
            print("\n⚠ WARNINGS:")
            for issue in issues:
                print(f"  - {issue}")
            return True
        else:
            print("\n✓ Split ratios are within expected ranges!")
            return True
    
    def _export_duplicates_to_excel(self, data, check_type, output_dir='results'):
        """Export duplicate data to Excel"""
        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"LEAKAGE_duplicate_{check_type}_report_{timestamp}.xlsx"
        
        # Create output directory if it doesn't exist
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        output_filepath = output_path / filename
        
        try:
            with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Duplicates', index=False)
                
                worksheet = writer.sheets['Duplicates']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"\n✓ Excel report exported: {output_filepath}")
            print(f"  Location: {output_filepath.absolute()}")
            print(f"  Total duplicate entries: {len(df)}")
            
        except Exception as e:
            print(f"\n✗ Error exporting: {e}")
    
    def run_complete_analysis(self, check_content=False, sample_size=None, output_dir='results'):
        """Run complete dataset analysis and validation"""
        print("\n" + "=" * 70)
        print("COMPREHENSIVE VIDEO DATASET ANALYSIS & VALIDATION")
        print("=" * 70)
        print(f"Split dataset: {self.split_path}")
        if self.original_path:
            print(f"Original dataset: {self.original_path}")
        print(f"Results will be saved to: {Path(output_dir).absolute()}")
        
        # Part 1: Structure Analysis
        print("\n" + "=" * 70)
        print("PART 1: DATASET STRUCTURE ANALYSIS")
        print("=" * 70)
        structure_df = self.export_structure_analysis(output_dir=output_dir)
        
        # Part 2: Leakage Detection
        print("\n" + "=" * 70)
        print("PART 2: DATA LEAKAGE DETECTION")
        print("=" * 70)
        
        results = {}
        results['filename'] = self.check_filename_duplicates(output_dir=output_dir)
        results['integrity'] = self.check_split_integrity()
        
        if check_content:
            results['content'] = self.check_content_duplicates(sample_size, output_dir=output_dir)
        
        # Final Summary
        print("\n" + "=" * 70)
        print("FINAL SUMMARY")
        print("=" * 70)
        
        if all(results.values()):
            print("\n✓✓✓ ALL CHECKS PASSED!")
            print("Dataset is properly structured with no data leakage detected.")
        else:
            print("\n✗✗✗ ISSUES DETECTED!")
            if not results.get('filename', True):
                print("→ Duplicate filenames found across splits")
            if not results.get('content', True):
                print("→ Identical content found across splits")
            if not results.get('integrity', True):
                print("→ Split distribution issues detected")
        
        print("=" * 70)
        
        return results


def main():
    # Configuration
    split_dataset_path = r"video_classification_project\data\raw"
    original_dataset_path = r"Dataset"  # Set to None if not available
    results_dir = "video_classification_project/results"  # Results directory
    
    print("=" * 70)
    print("VIDEO DATASET ANALYZER & VALIDATOR")
    print("=" * 70)
    print(f"Working directory: {os.getcwd()}")
    print(f"Split dataset: {Path(split_dataset_path).absolute()}")
    if original_dataset_path:
        print(f"Original dataset: {Path(original_dataset_path).absolute()}")
    print(f"Results directory: {Path(results_dir).absolute()}")
    
    try:
        analyzer = VideoDatasetAnalyzer(split_dataset_path, original_dataset_path)
        
        # Run complete analysis
        analyzer.run_complete_analysis(check_content=False, output_dir=results_dir)
        
        # Optional: Run content-based check
        print("\n" + "=" * 70)
        print("OPTIONAL: CONTENT-BASED DUPLICATE CHECK")
        print("=" * 70)
        print("This uses file hashing to detect identical videos with different names.")
        print("WARNING: Can be slow for large datasets.")
        print("\nOptions:")
        print("  1. Skip")
        print("  2. Full check (all files)")
        print("  3. Sample check (quick)")
        
        choice = input("\nEnter choice (1-3) [1]: ").strip() or "1"
        
        if choice == '2':
            analyzer.check_content_duplicates(sample_size=None, output_dir=results_dir)
        elif choice == '3':
            sample = int(input("Sample size per split [100]: ").strip() or "100")
            analyzer.check_content_duplicates(sample_size=sample, output_dir=results_dir)
        
        print("\n✓ Analysis complete!")
        
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()