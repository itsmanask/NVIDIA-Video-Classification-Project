import os
from pathlib import Path
from collections import defaultdict
import pandas as pd

def analyze_original_dataset(original_path, category_structure, category_mapping):
    """
    Analyze original dataset structure: original_path/category/videos/subcategory
    Returns dictionary with counts per category/subcategory
    """
    original_dir = Path(original_path)
    video_extensions = {'.mp4', '.avi', '.mov', '.mkv'}
    original_counts = {}
    
    if not original_dir.exists():
        print(f"Warning: Original dataset path {original_path} does not exist!")
        return original_counts
    
    print(f"\nAnalyzing original dataset at: {original_path}")
    
    for main_category, subcategories in category_structure.items():
        # Use the original dataset folder name (with spaces)
        original_category_name = category_mapping.get(main_category, main_category)
        
        for subcategory in subcategories:
            # Updated path: original_path/category/videos/subcategory
            subcategory_path = original_dir / original_category_name / 'videos' / subcategory
            
            video_count = 0
            if subcategory_path.exists():
                all_files = list(subcategory_path.iterdir())
                video_files = [
                    f for f in all_files 
                    if f.is_file() and f.suffix.lower() in video_extensions
                ]
                video_count = len(video_files)
            
            original_counts[(main_category, subcategory)] = video_count
    
    return original_counts

def analyze_and_export_dataset(base_path, original_path=None, output_file='dataset_split_analysis.xlsx'):
    """
    Analyze dataset structure with train/val/test splits and original dataset
    Structure: 
    - Split: base_path/[train|val|test]/categories/subcategories
    - Original: original_path/category/subcategory
    """
    base_dir = Path(base_path)
    
    if not base_dir.exists():
        print(f"Directory {base_path} does not exist!")
        return
    
    print(f"Analyzing split dataset at: {base_path}")
    print("=" * 80)
    
    # Mapping from split dataset names (with underscores) to original dataset names (with spaces)
    category_mapping = {
        'Natural_Content': 'Natural Content',
        'Flat_Content': 'Flat Content',
        'Animation': 'Animation',
        'Gaming': 'Gaming'
    }
    
    # Define all subcategories for each main category
    # Keys use underscore format (for split dataset)
    category_structure = {
        'Animation': [
            'Animation', 'Bleach', 'Cartoon', 'Dragon Ball',
            'Lego minifigure', 'Naruto',
            'One Piece', 'Sonic the Hedgehog',
            'The Walt Disney Company'
        ],
        
        'Gaming': [
            'Action-adventure game', 'Battlefield', 'Call of Duty',
            'FIFA 15', 'Games',
            'Grand Theft Auto', 'Grand Theft Auto V',
            'League of Legends', 'Minecraft',
            'RuneScape', 'Video game', 'World of Warcraft'
        ],
        
        'Natural_Content': [
            'Animal', 'Bird', 'Cat', 'Chicken',
            'Dog', 'Farm', 'Fish',
            'Fishing', 'Garden', 'Horse',
            'Nature', 'Outdoor recreation', 'Pet',
            'Plant', 'Tree', 'Wildlife'
        ],
        
        'Flat_Content': [
            'Chart', 'Illustration', 'Logo', 'Map',
            'Poster', 'Screencast', 'Text',
            'Typography', 'Website'
        ]
    }
    
    # Analyze original dataset if path provided
    original_counts = {}
    if original_path:
        original_counts = analyze_original_dataset(original_path, category_structure, category_mapping)
    
    splits = ['train', 'val', 'test']
    video_extensions = {'.mp4', '.avi', '.mov', '.mkv'}
    
    # Storage for counts
    data_rows = []
    
    # Analyze each category and subcategory
    for main_category, subcategories in category_structure.items():
        print(f"\nProcessing {main_category}...")
        
        for subcategory in subcategories:
            row_data = {
                'Main Category': main_category,
                'Subcategory': subcategory,
                'Original': original_counts.get((main_category, subcategory), 0) if original_path else 0,
                'Train': 0,
                'Validation': 0,
                'Test': 0,
                'Total Split': 0
            }
            
            for split in splits:
                split_path = base_dir / split
                
                if not split_path.exists():
                    continue
                
                # Path: base_path/[split]/[category]/[subcategory]
                subcategory_path = split_path / main_category / subcategory
                
                video_count = 0
                if subcategory_path.exists():
                    all_files = list(subcategory_path.iterdir())
                    video_files = [
                        f for f in all_files 
                        if f.is_file() and f.suffix.lower() in video_extensions
                    ]
                    video_count = len(video_files)
                
                # Map split names to column names
                split_column_map = {
                    'train': 'Train',
                    'val': 'Validation',
                    'test': 'Test'
                }
                
                row_data[split_column_map[split]] = video_count
                row_data['Total Split'] += video_count
            
            data_rows.append(row_data)
            if original_path:
                print(f"  {subcategory}: Original={row_data['Original']}, Train={row_data['Train']}, Val={row_data['Validation']}, Test={row_data['Test']}, Split Total={row_data['Total Split']}")
            else:
                print(f"  {subcategory}: Train={row_data['Train']}, Val={row_data['Validation']}, Test={row_data['Test']}, Total={row_data['Total Split']}")
    
    # Create DataFrame
    df = pd.DataFrame(data_rows)
    
    # Sort by Main Category and Subcategory name (ascending alphabetically)
    df = df.sort_values(['Main Category', 'Subcategory'], ascending=[True, True])
    
    # Calculate category totals
    category_totals = []
    for category in category_structure.keys():
        cat_df = df[df['Main Category'] == category]
        total_row = {
            'Main Category': '',
            'Subcategory': f'Total: {len(cat_df)}',
            'Original': cat_df['Original'].sum() if original_path else 0,
            'Train': cat_df['Train'].sum(),
            'Validation': cat_df['Validation'].sum(),
            'Test': cat_df['Test'].sum(),
            'Total Split': cat_df['Total Split'].sum()
        }
        category_totals.append((category, total_row))
    
    # Insert total rows after each category
    final_rows = []
    for category in category_structure.keys():
        # Add all rows for this category
        cat_rows = df[df['Main Category'] == category].to_dict('records')
        final_rows.extend(cat_rows)
        
        # Add total row
        total_row = [t for c, t in category_totals if c == category][0]
        final_rows.append(total_row)
        
        # Add blank row for separation (except after last category)
        if category != list(category_structure.keys())[-1]:
            final_rows.append({
                'Main Category': '',
                'Subcategory': '',
                'Original': '',
                'Train': '',
                'Validation': '',
                'Test': '',
                'Total Split': ''
            })
    
    # Create final DataFrame
    final_df = pd.DataFrame(final_rows)
    
    # Export to Excel with formatting
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name='Dataset Analysis', index=False)
            
            # Get the worksheet
            worksheet = writer.sheets['Dataset Analysis']
            
            # Format header row
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color='B4C7DC', end_color='B4C7DC', fill_type='solid')
            header_font = Font(bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 12
            worksheet.column_dimensions['D'].width = 12
            worksheet.column_dimensions['E'].width = 12
            worksheet.column_dimensions['F'].width = 12
            worksheet.column_dimensions['G'].width = 12
            
            # Format total rows (bold and highlighted)
            total_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            total_font = Font(bold=True)
            
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                cell_value = worksheet[f'B{idx}'].value
                if cell_value and str(cell_value).startswith('Total:'):
                    for cell in row:
                        cell.fill = total_fill
                        cell.font = total_font
            
            # Center align numeric columns
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                          min_col=3, max_col=7):
                for cell in row:
                    if cell.value != '':
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"\n✓ Successfully exported to: {output_file}")
        print(f"  Location: {Path(output_file).absolute()}")
        
        # Print summary statistics
        print("\n" + "=" * 80)
        print("SUMMARY STATISTICS")
        print("=" * 80)
        
        if original_path:
            total_original = df['Original'].sum()
            print(f"\nOriginal Dataset: {total_original} videos")
        
        total_videos = df['Total Split'].sum()
        total_train = df['Train'].sum()
        total_val = df['Validation'].sum()
        total_test = df['Test'].sum()
        
        print(f"\nSplit Dataset Total: {total_videos} videos")
        print(f"  Train:      {total_train:5} videos ({total_train/total_videos*100:5.1f}%)")
        print(f"  Validation: {total_val:5} videos ({total_val/total_videos*100:5.1f}%)")
        print(f"  Test:       {total_test:5} videos ({total_test/total_videos*100:5.1f}%)")
        
        print("\nCategory Breakdown:")
        for category in category_structure.keys():
            cat_df = df[df['Main Category'] == category]
            if original_path:
                cat_original = cat_df['Original'].sum()
                cat_split = cat_df['Total Split'].sum()
                print(f"  {category:20}: Original={cat_original:5} | Split={cat_split:5} ({cat_split/total_videos*100:5.1f}%)")
            else:
                cat_total = cat_df['Total Split'].sum()
                print(f"  {category:20}: {cat_total:5} videos ({cat_total/total_videos*100:5.1f}%)")
        
        return final_df
        
    except Exception as e:
        print(f"\n✗ Error exporting to Excel: {e}")
        print("\nMake sure you have openpyxl installed:")
        print("  pip install openpyxl pandas")
        
        # Fallback: export to CSV
        csv_file = output_file.replace('.xlsx', '.csv')
        final_df.to_csv(csv_file, index=False)
        print(f"\n✓ Exported to CSV instead: {csv_file}")
        
        return final_df

if __name__ == "__main__":
    # Update these paths to your dataset locations
    split_dataset_path = r"video_classification_project\data\raw"
    original_dataset_path = r"Dataset"  # Update this to your original dataset path
    output_file = "dataset_complete_analysis.xlsx"
    
    print("Dataset Analysis - Original and Split Datasets")
    print("===============================================")
    print(f"Current working directory: {os.getcwd()}")
    print(f"Split dataset at: {Path(split_dataset_path).absolute()}")
    print(f"Original dataset at: {Path(original_dataset_path).absolute()}")
    print()
    
    try:
        df = analyze_and_export_dataset(
            base_path=split_dataset_path,
            original_path=original_dataset_path,
            output_file=output_file
        )
        
        if df is not None:
            print("\n✓ Analysis complete!")
            print(f"\nYou can now open: {output_file}")
        else:
            print("\n⚠ Analysis completed but no data exported.")
            
    except Exception as e:
        print(f"Error analyzing dataset: {e}")
        import traceback
        traceback.print_exc()
        
        print("\nPlease ensure:")
        print("1. Both dataset paths are correct")
        print("2. Split directory structure: base_path/[train|val|test]/category/subcategory/")
        print("3. Original directory structure: original_path/category/videos/subcategory/")
        print("4. You have pandas and openpyxl installed:")
        print("   pip install pandas openpyxl")